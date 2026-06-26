import streamlit as st
import pandas as pd
from datetime import datetime
import numpy as np
import re
from io import BytesIO

st.set_page_config(
    page_title="DSP Delivery Tracker",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)
st.markdown(
    '<style>'
    '.stDeployButton{display:none!important}'
    '</style>',
    unsafe_allow_html=True
)
TODAY = datetime.today()
CURRENT_MONTH = TODAY.month
CURRENT_YEAR = TODAY.year

with st.sidebar:
    st.title("📊 DSP Delivery Tracker")
    st.markdown("---")
    st.subheader("📝 Report Name")
    report_name = st.text_input(
        "Name this report",
        value="",
        placeholder="e.g. June Week 3"
    )
    st.markdown("---")
    st.subheader("📁 Upload Files")
    uploaded_file = st.file_uploader(
        "1. Overall Data (Full YTD-MTD)",
        type=['csv', 'xlsx', 'xls']
    )
    uploaded_3day = st.file_uploader(
        "2. Last 3 Days Data (for DRR)",
        type=['csv', 'xlsx', 'xls']
    )
    st.markdown("---")
    st.subheader("📅 Projection")
    projection_date = st.date_input(
        "Projection Date",
        datetime(2026, 6, 30)
    )
    st.markdown("---")
    st.subheader("⚙️ Thresholds")
    under_threshold = st.slider(
        "Under below (%)", 80, 100, 98
    )
    over_threshold = st.slider(
        "Over above (%)", 100, 120, 105
    )
    st.markdown("---")
    show_only_delivering = st.checkbox(
        "Show only Delivering", value=True
    )


def extract_budget(name):
    if pd.isna(name):
        return 0
    matches = []
    for m in re.finditer(
        r'(\d+\.?\d*)\s*([KkLl])', str(name)
    ):
        val = float(m.group(1))
        u = m.group(2).upper()
        if u == 'L':
            matches.append((m.start(), val * 100000))
        else:
            matches.append((m.start(), val * 1000))
    if matches:
        matches.sort(key=lambda x: x[0])
        return matches[-1][1]
    return 0


def parse_file(f):
    if f.name.endswith('.csv'):
        return pd.read_csv(f)
    return pd.read_excel(f)


def std_cols(df):
    cmap = {}
    used = set()
    for col in df.columns:
        cl = col.lower().strip()
        m = None
        if 'order status' in cl:
            m = 'Order Status'
        elif 'campaign name' in cl:
            m = 'Order Name'
        elif 'advertiser' in cl and 'account' in cl:
            m = 'Account'
        elif 'start' in cl and 'date' in cl:
            m = 'Start Date'
        elif 'end' in cl and 'date' in cl:
            m = 'End Date'
        elif 'budget' in cl:
            m = 'Budget'
        elif 'total cost' in cl:
            m = 'Total Spend'
        elif cl == 'impressions':
            m = 'Impressions'
        elif 'click' in cl:
            m = 'Clicks'
        elif cl == 'ctr':
            m = 'CTR'
        elif 'total roas' in cl and 'click' not in cl:
            m = 'ROAS'
        elif 'total dpvr' in cl:
            m = 'DPVR'
        elif 'total purchases' in cl:
            m = 'Purchases'
        elif 'ecpm' in cl:
            m = 'eCPM'
        elif 'new-to-brand' in cl or 'ntb' in cl:
            m = 'NTB'
        if m and m not in used:
            cmap[col] = m
            used.add(m)
    df = df.rename(columns=cmap)
    return df.loc[:, ~df.columns.duplicated(keep='first')]


def process(df, today, drr_data=None, proj_date=None):
    df = df.copy().dropna(how='all')
    df = std_cols(df)
    if 'Order Name' not in df.columns:
        st.error("No Campaign name column found.")
        return pd.DataFrame()
    df = df.dropna(subset=['Order Name'])
    df = df[df['Order Name'].str.strip() != '']
    df['Start Date'] = pd.to_datetime(
        df['Start Date'], errors='coerce'
    )
    df['End Date'] = pd.to_datetime(
        df['End Date'], errors='coerce'
    )
    today = pd.Timestamp(today)
    for c in ['Budget', 'Total Spend', 'Impressions',
              'Clicks', 'CTR', 'ROAS', 'DPVR',
              'Purchases', 'eCPM', 'NTB']:
        if c in df.columns:
            df[c] = pd.to_numeric(
                df[c], errors='coerce'
            ).fillna(0)
    if 'NTB' not in df.columns:
        df['NTB'] = 0
    df['Budget'] = df.apply(
        lambda r: extract_budget(r['Order Name'])
        if r.get('Budget', 0) == 0
        else r['Budget'], axis=1
    )
    df['Total Days'] = (
        df['End Date'] - df['Start Date']
    ).dt.days + 1
    df['Elapsed Days'] = (
        (today - df['Start Date']).dt.days + 1
    ).clip(lower=0)
    df['Elapsed Days'] = df[
        ['Elapsed Days', 'Total Days']
    ].min(axis=1)
    df['Remaining Days'] = (
        df['Total Days'] - df['Elapsed Days']
    ).clip(lower=0)
    df['Daily Budget'] = np.where(
        df['Total Days'] > 0,
        df['Budget'] / df['Total Days'], 0
    )
    df['Ideal Spend'] = (
        df['Daily Budget'] * df['Elapsed Days']
    )
    df['Remaining Budget'] = (
        df['Budget'] - df['Total Spend']
    ).clip(lower=0)
    df['DR %'] = np.where(
        df['Ideal Spend'] > 0,
        np.round(
            df['Total Spend'] / df['Ideal Spend'] * 100,
            1
        ), 0
    )
    df['Expected DRR'] = np.where(
        df['Remaining Days'] > 0,
        df['Remaining Budget'] / df['Remaining Days'],
        0
    )
    df['Current DRR'] = 0.0
    if drr_data is not None:
        dd = std_cols(drr_data.copy())
        if 'Order Name' in dd.columns:
            if 'Total Spend' in dd.columns:
                dd['Total Spend'] = pd.to_numeric(
                    dd['Total Spend'], errors='coerce'
                ).fillna(0)
                dl = dd.groupby('Order Name')[
                    'Total Spend'
                ].sum().reset_index()
                dl.columns = ['Order Name', 'S3']
                dl['Current DRR'] = dl['S3'] / 3
                df = df.merge(
                    dl[['Order Name', 'Current DRR']],
                    on='Order Name', how='left',
                    suffixes=('_x', '')
                )
                if 'Current DRR_x' in df.columns:
                    df['Current DRR'] = df[
                        'Current DRR'
                    ].fillna(df['Current DRR_x'])
                    df.drop(
                        'Current DRR_x',
                        axis=1, inplace=True
                    )
                df['Current DRR'] = (
                    df['Current DRR'].fillna(0)
                )
    else:
        df['Current DRR'] = np.where(
            df['Elapsed Days'] > 0,
            df['Total Spend'] / df['Elapsed Days'], 0
        )
    edrr = np.where(
        df['Current DRR'] > 0,
        df['Current DRR'],
        df['Daily Budget']
    )
    df['Projected at End'] = (
        df['Total Spend'] + edrr * df['Remaining Days']
    )
    df['Projected at End'] = df[
        ['Projected at End', 'Budget']
    ].min(axis=1)
    df['Proj End %'] = np.where(
        df['Budget'] > 0,
        np.round(
            df['Projected at End'] / df['Budget'] * 100,
            1
        ), 0
    )
    df['Projected Spend'] = df['Total Spend']
    if proj_date is not None:
        pts = pd.Timestamp(proj_date)
        if pts > today:
            dp = (pts - today).days
            df['Projected Spend'] = (
                df['Total Spend'] + edrr * dp
            )
            df['Projected Spend'] = df[
                ['Projected Spend', 'Budget']
            ].min(axis=1)

    def status(r):
        os = r.get('Order Status', '')
        if os == 'Ended':
            return 'Ended'
        if os == 'Inactive':
            return 'Inactive'
        if os == 'Line items not running':
            return 'Not Spending'
        if r['Budget'] == 0:
            return 'No Budget'
        if r['Total Spend'] == 0 and r['Elapsed Days'] > 3:
            return 'Not Spending'
        pe = r['Proj End %']
        if pe >= under_threshold and pe <= over_threshold:
            return 'On Track'
        elif pe > over_threshold:
            return 'Over-delivering'
        return 'Under-delivering'

    df['Status'] = df.apply(status, axis=1)
    df['Account Short'] = df['Account'].str.replace(
        'IN - GCS - CEPC - ', '', regex=False
    ).str.strip()
    df['CTR %'] = df['CTR'] * 100
    df['DPVR %'] = df['DPVR'] * 100
    return df


ICONS = {
    'On Track': '🟢',
    'Under-delivering': '🟡',
    'Over-delivering': '🔵',
    'Not Spending': '🔴',
    'No Budget': '⚪',
    'Inactive': '⚪',
    'Ended': '⏹️'
}


def b_ctr(v):
    try:
        v = float(v)
        if v > 0.60:
            return '🟢 ' + str(round(v, 2)) + '%'
        elif v >= 0.40:
            return '🟡 ' + str(round(v, 2)) + '%'
        return '🔴 ' + str(round(v, 2)) + '%'
    except Exception:
        return str(v)


def b_dpvr(d, c):
    try:
        d, c = float(d), float(c)
        if d > c:
            return '🟢 ' + str(round(d, 2)) + '%'
        return '🔴 ' + str(round(d, 2)) + '%'
    except Exception:
        return str(d)


def b_ntb(v):
    try:
        v = float(v)
        if v > 0.6:
            return '🟢 ' + str(round(v*100, 1)) + '%'
        elif v >= 0.4:
            return '🟡 ' + str(round(v*100, 1)) + '%'
        return '🔴 ' + str(round(v*100, 1)) + '%'
    except Exception:
        return str(v)


def b_roas(v):
    try:
        v = float(v)
        if v > 2:
            return '🟢 ' + str(round(v, 2))
        elif v >= 1:
            return '🟡 ' + str(round(v, 2))
        return '🔴 ' + str(round(v, 2))
    except Exception:
        return str(v)


def make_excel(acct_df, order_df, perf_df, info):
    out = BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        pd.DataFrame([info]).to_excel(
            w, sheet_name='Summary', index=False
        )
        if len(acct_df) > 0:
            acct_df.to_excel(
                w, sheet_name='Accounts', index=False
            )
        if len(order_df) > 0:
            order_df.to_excel(
                w, sheet_name='Orders', index=False
            )
        if len(perf_df) > 0:
            perf_df.to_excel(
                w, sheet_name='Performance', index=False
            )
        from openpyxl.styles import Font, PatternFill
        hf = Font(bold=True, color='FFFFFF')
        hfill = PatternFill(
            start_color='232F3E',
            end_color='232F3E',
            fill_type='solid'
        )
        for sn in w.book.sheetnames:
            ws = w.book[sn]
            for cell in ws[1]:
                cell.font = hf
                cell.fill = hfill
    out.seek(0)
    return out


if uploaded_file is not None:
    raw = parse_file(uploaded_file)
    drr = parse_file(uploaded_3day) if uploaded_3day else None
    df = process(raw, TODAY, drr_data=drr, proj_date=projection_date)
    if len(df) == 0:
        st.error("No data.")
        st.stop()
    if show_only_delivering:
        adf = df[df['Order Status'] == 'Delivering'].copy()
    else:
        adf = df[~df['Status'].isin(
            ['Ended', 'Inactive']
        )].copy()
    acm = df[
        (df['Start Date'] <= pd.Timestamp(
            CURRENT_YEAR, CURRENT_MONTH, 28
        ))
        & (df['End Date'] >= pd.Timestamp(
            CURRENT_YEAR, CURRENT_MONTH, 1
        ))
    ].copy()
    afd = df[df['Order Status'].isin(
        ['Delivering', 'Inactive',
         'Line items not running']
    )]
    fs = afd['Start Date'].min()
    fe = afd['End Date'].max()
    dc = len(df[df['Order Status'] == 'Delivering'])
    ec = len(df[df['Order Status'] == 'Ended'])
    ic = (
        len(df[df['Order Status'] == 'Inactive'])
        + len(df[
            df['Order Status'] == 'Line items not running'
        ])
    )
    ta = df['Account Short'].nunique()
    aa = adf['Account Short'].nunique()
    ea = df[
        df['Order Status'] == 'Ended'
    ]['Account Short'].nunique()
    pts = pd.Timestamp(projection_date)
    tts = pd.Timestamp(TODAY)
    sp = pts > tts

    title = "📊 DSP Delivery Tracker"
    if report_name:
        title += " - " + report_name
    st.title(title)
    fss = fs.strftime('%d %b %Y') if pd.notna(fs) else ''
    fes = fe.strftime('%d %b %Y') if pd.notna(fe) else ''
    st.caption(
        TODAY.strftime('%d %B %Y')
        + " | Flight: " + fss + " to " + fes
        + " | Orders: " + str(len(df))
    )
    if uploaded_3day:
        st.success("3-Day DRR loaded")
    else:
        st.info("Upload Last 3 Days for accurate DRR")
    st.markdown("---")

    # ACCOUNT SUMMARY
    st.subheader("🏢 Account Level Summary")
    asf = st.multiselect(
        "Select Accounts",
        sorted(adf['Account Short'].unique()),
        default=[], key="as1"
    )
    sd = adf.copy()
    if asf:
        sd = sd[sd['Account Short'].isin(asf)]
    sb = sd[sd['Budget'] > 0]['Budget'].sum()
    si = sd[sd['Budget'] > 0]['Ideal Spend'].sum()
    ss2 = sd['Total Spend'].sum()
    sdr = (ss2 / si * 100) if si > 0 else 0
    sag = sd[sd['Budget'] > 0].groupby(
        'Account Short'
    ).agg({
        'Budget': 'sum', 'Projected at End': 'sum'
    }).reset_index()
    sag['P'] = np.where(
        sag['Budget'] > 0,
        sag['Projected at End'] / sag['Budget'] * 100, 0
    )
    su = len(sag[sag['P'] < under_threshold])
    so = len(sag[sag['P'] > over_threshold])
    sn = len(sag) - su - so
    c1, c2, c3, c4, c5, c6, c7, c8 = st.columns(8)
    c1.metric("Total", ta)
    c2.metric("Active", len(sag))
    c3.metric("Ended", ea)
    c4.metric("🟡Under", su)
    c5.metric("🔵Over", so)
    c6.metric("🟢OnTrack", sn)
    c7.metric("Budget", "₹" + str(round(sb/100000, 1)) + "L")
    c8.metric("DR%", str(round(sdr, 1)) + "%")
    st.markdown("---")

    # ORDER SUMMARY
    st.subheader("📋 Order Level Summary")
    sc = adf['Status'].value_counts()
    tb = adf[adf['Budget'] > 0]['Budget'].sum()
    ard = adf[
        (adf['Proj End %'] < 80) & (adf['Budget'] > 0)
    ]
    ar = ard['Remaining Budget'].sum()
    d1, d2, d3, d4, d5, d6, d7, d8 = st.columns(8)
    d1.metric("Total", len(df))
    d2.metric("🟢Active", dc)
    d3.metric("⏹️Ended", ec)
    d4.metric("⚪Inactive", ic)
    d5.metric("🟢OnTrack", sc.get('On Track', 0))
    d6.metric("🟡Under", sc.get('Under-delivering', 0))
    d7.metric("🔵Over", sc.get('Over-delivering', 0))
    d8.metric("⚠️Risk", "₹" + str(round(ar/100000, 1)) + "L")
    if sp:
        st.markdown("---")
        pt = adf['Projected Spend'].sum()
        st.subheader(
            "🔮 Projection by "
            + projection_date.strftime('%d %b %Y')
        )
        e1, e2, e3 = st.columns(3)
        e1.metric("Days", str((pts - tts).days))
        e2.metric(
            "Projected",
            "₹" + str(round(pt/100000, 1)) + "L"
        )
        e3.metric(
            "Remaining",
            "₹" + str(round((tb-pt)/100000, 1)) + "L"
        )
    st.markdown("---")

    # ACCOUNT TABLE
    st.header("🏢 Account-Level Overview")
    st.caption("All orders in " + TODAY.strftime('%B %Y'))
    aod = acm.copy()
    ad = pd.DataFrame()
    if len(aod) > 0:
        atf = st.multiselect(
            "Filter Account",
            sorted(aod['Account Short'].unique()),
            default=[], key="at1"
        )
        if atf:
            aod = aod[aod['Account Short'].isin(atf)]
        acs = aod.groupby('Account Short').agg({
            'Budget': 'sum',
            'Total Spend': 'sum',
            'Ideal Spend': 'sum',
            'Current DRR': 'sum',
            'Expected DRR': 'sum',
            'Projected Spend': 'sum',
            'Projected at End': 'sum',
            'Start Date': 'min',
            'End Date': 'max',
            'CTR %': 'mean',
            'DPVR %': 'mean',
            'ROAS': 'mean',
            'Order Name': 'count'
        }).reset_index()
        acs.columns = [
            'Account', 'Budget', 'Spends',
            'Ideal Spend', 'Current DRR',
            'Expected DRR', 'Projected Spend',
            'Proj End', 'Start Date', 'End Date',
            'CTR %', 'DPVR %', 'ROAS', 'Orders'
        ]
        acs['DR %'] = np.where(
            acs['Ideal Spend'] > 0,
            np.round(
                acs['Spends'] / acs['Ideal Spend'] * 100,
                1
            ), 0
        )

        def ast(r):
            if r['Budget'] <= 0:
                return 'Under-delivering'
            p = r['Proj End'] / r['Budget'] * 100
            if p >= under_threshold and p <= over_threshold:
                return 'On Track'
            elif p > over_threshold:
                return 'Over-delivering'
            return 'Under-delivering'

        acs['Status'] = acs.apply(ast, axis=1)
        acs = acs.sort_values('DR %')
        acs['Start Date'] = acs[
            'Start Date'
        ].dt.strftime('%d %b %Y')
        acs['End Date'] = acs[
            'End Date'
        ].dt.strftime('%d %b %Y')
        cols = [
            'Account', 'Start Date', 'End Date',
            'Budget', 'Spends', 'DR %',
            'Current DRR', 'Expected DRR'
        ]
        if sp:
            cols.append('Projected Spend')
        cols += ['CTR %', 'DPVR %', 'ROAS',
                 'Orders', 'Status']
        ad = acs[cols].copy()
        ad['Status'] = ad['Status'].map(
            lambda x: ICONS.get(x, '') + " " + x
        )
        fm = {
            'Budget': '₹{:,.0f}',
            'Spends': '₹{:,.0f}',
            'DR %': '{:.1f}%',
            'Current DRR': '₹{:,.0f}',
            'Expected DRR': '₹{:,.0f}',
            'CTR %': '{:.2f}%',
            'DPVR %': '{:.2f}%',
            'ROAS': '{:.2f}'
        }
        if sp:
            fm['Projected Spend'] = '₹{:,.0f}'
        st.dataframe(
            ad.style.format(fm),
            use_container_width=True,
            height=min(650, max(250, len(acs) * 38))
        )
    st.markdown("---")

    # ORDER TABLE
    st.header("📋 Order-Level Tracker")
    f1, f2, f3 = st.columns([2, 2, 1])
    with f1:
        oa = st.multiselect(
            "Filter Account",
            sorted(adf['Account Short'].unique()),
            default=[], key="o1"
        )
    with f2:
        os2 = st.multiselect(
            "Filter Status",
            adf['Status'].unique().tolist(),
            default=adf['Status'].unique().tolist()
        )
    with f3:
        pr = st.slider("DR%", 0, 200, (0, 200))
    fd = adf.copy()
    if oa:
        fd = fd[fd['Account Short'].isin(oa)]
    fd = fd[
        (fd['Status'].isin(os2))
        & (fd['DR %'].between(pr[0], pr[1]))
    ]
    odata = {
        'Account': fd['Account Short'].values,
        'Order': fd['Order Name'].values,
        'Start': fd['Start Date'].dt.strftime(
            '%d %b %Y'
        ).values,
        'End': fd['End Date'].dt.strftime(
            '%d %b %Y'
        ).values,
        'Budget': fd['Budget'].values,
        'Spends': fd['Total Spend'].values,
        'DR %': fd['DR %'].values,
        'Cur DRR': fd['Current DRR'].values,
        'Exp DRR': fd['Expected DRR'].values,
    }
    if sp:
        odata['Proj Spend'] = fd['Projected Spend'].values
    odata['CTR %'] = fd['CTR %'].values
    odata['DPVR %'] = fd['DPVR %'].values
    odata['ROAS'] = fd['ROAS'].values
    odata['Status'] = [
        ICONS.get(s, '') + " " + s
        for s in fd['Status'].values
    ]
    odf = pd.DataFrame(odata).sort_values('DR %')
    ofm = {
        'Budget': '₹{:,.0f}',
        'Spends': '₹{:,.0f}',
        'DR %': '{:.1f}%',
        'Cur DRR': '₹{:,.0f}',
        'Exp DRR': '₹{:,.0f}',
        'CTR %': '{:.2f}%',
        'DPVR %': '{:.2f}%',
        'ROAS': '{:.2f}'
    }
    if sp:
        ofm['Proj Spend'] = '₹{:,.0f}'
    st.dataframe(
        odf.style.format(ofm),
        use_container_width=True,
        height=600
    )
    st.markdown("---")

    # PERFORMANCE
    st.header("📈 Performance")
    st.caption("🟢Good 🟡Average 🔴Poor")
    pv = st.radio("View:", ["Account", "Order"], horizontal=True)
    pdf = pd.DataFrame()
    if pv == "Account":
        ps = adf[adf['Budget'] > 0]
        if len(ps) > 0:
            pa = ps.groupby('Account Short').agg({
                'CTR %': 'mean',
                'DPVR %': 'mean',
                'NTB': 'mean',
                'ROAS': 'mean',
                'Total Spend': 'sum',
                'Order Name': 'count'
            }).reset_index()
            pa.columns = [
                'Account', 'CTR %', 'DPVR %',
                'NTB', 'ROAS', 'Spend', 'Orders'
            ]
            pa = pa.sort_values('Spend', ascending=False)
            pdf = pd.DataFrame({
                'Account': pa['Account'].values,
                'CTR': [b_ctr(v) for v in pa['CTR %']],
                'DPVR': [
                    b_dpvr(d, c) for d, c
                    in zip(pa['DPVR %'], pa['CTR %'])
                ],
                'NTB': [b_ntb(v) for v in pa['NTB']],
                'ROAS': [b_roas(v) for v in pa['ROAS']],
                'Orders': pa['Orders'].values
            })
            st.dataframe(pdf, use_container_width=True)
    else:
        pa2 = st.selectbox(
            "Account",
            ["All"] + sorted(
                adf['Account Short'].unique().tolist()
            )
        )
        po = adf[adf['Budget'] > 0].copy()
        if pa2 != "All":
            po = po[po['Account Short'] == pa2]
        if len(po) > 0:
            pdf = pd.DataFrame({
                'Account': po['Account Short'].values,
                'Order': po['Order Name'].str[:50].values,
                'CTR': [b_ctr(v) for v in po['CTR %']],
                'DPVR': [
                    b_dpvr(d, c) for d, c
                    in zip(po['DPVR %'], po['CTR %'])
                ],
                'NTB': [b_ntb(v) for v in po['NTB']],
                'ROAS': [b_roas(v) for v in po['ROAS']],
            })
            st.dataframe(pdf, use_container_width=True)
    st.markdown("---")

    # DOWNLOAD
    st.header("📥 Download")
    nm = report_name if report_name else "DSP_Tracker"
    x1, x2 = st.columns(2)
    with x1:
        info = {
            'Report': nm,
            'Date': TODAY.strftime('%d %b %Y'),
            'Orders': len(df),
            'Active': dc,
            'Budget': sb,
            'Spend': ss2
        }
        xl = make_excel(ad, odf, pdf, info)
        st.download_button(
            "📊 Download Excel Report",
            xl,
            nm + ".xlsx",
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    with x2:
        st.download_button(
            "📄 Download CSV",
            adf.to_csv(index=False),
            nm + ".csv",
            "text/csv"
        )

else:
    st.title("📊 DSP Delivery Tracker")
    st.markdown("### Upload Entity Order Summary")
    st.markdown("---")
    st.markdown(
        "**File 1:** Entity Order Summary "
        "(Full YTD-MTD)\n\n"
        "**File 2:** Last 3 Days (for DRR)"
    )
    st.markdown("---")
    st.markdown(
        "| Status | Condition |\n"
        "|---|---|\n"
        "| 🟢 On Track | Proj end 98-105% |\n"
        "| 🟡 Under | Proj end < 98% |\n"
        "| 🔵 Over | Proj end > 105% |\n"
        "| 🔴 Not Spending | Zero 3+ days |"
    )
    st.info("Upload file in sidebar to begin!")
